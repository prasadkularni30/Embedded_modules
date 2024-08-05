import paho.mqtt.client as mqtt
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json
import os
from datetime import datetime
import config
import logging

# Set up logging
logging.basicConfig(level=logging.DEBUG)

# Create a new Excel file or load existing one
if os.path.exists(config.EXCEL_FILE):
    workbook = openpyxl.load_workbook(config.EXCEL_FILE)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    # Set up headers
    headers = ["Timestamp", "Temperature", "Humidity"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

# MQTT Callbacks
def on_connect(client, userdata, flags, reason_code,properties):
    if reason_code == 0:
        logging.info("Connected to MQTT Broker successfully")
        client.subscribe(config.MQTT_TOPIC)
    else:
        logging.error(f"Failed to connect, return code {reason_code}")

def on_message(client, userdata, msg):
    try:
        # Parse the incoming message
        payload = json.loads(msg.payload.decode())
        logging.info(f"Received message: {payload}")

        # Extract data
        temperature = payload.get("temperature", "N/A")
        humidity = payload.get("humidity", "N/A")
        
        # Get current timestamp
        timestamp = datetime.now()

        # Append data to Excel
        next_row = sheet.max_row + 1
        sheet[f"A{next_row}"] = timestamp
        sheet[f"B{next_row}"] = temperature
        sheet[f"C{next_row}"] = humidity

        # Save the Excel file
        workbook.save(config.EXCEL_FILE)
        logging.info("Data saved to Excel")
    except Exception as e:
        logging.error(f"Error processing message: {e}")

# Initialize MQTT Client
client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2,client_id=config.MQTT_CLIENT_ID_SUBSCRIBER)

client.username_pw_set(config.MQTT_USERNAME, config.MQTT_PASSWORD)
client.on_connect = on_connect
client.on_message = on_message

# Enable detailed logging for the MQTT client
client.enable_logger(logging.getLogger())

# Connect to MQTT Broker
try:
    logging.info(f"Connecting to MQTT Broker at {config.MQTT_BROKER}:{config.MQTT_PORT}")
    client.connect(config.MQTT_BROKER, config.MQTT_PORT, 60)
except Exception as e:
    logging.error(f"Failed to connect to MQTT Broker: {e}")

# Start the MQTT loop
client.loop_forever()